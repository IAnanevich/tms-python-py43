# Прочитать сохраненный csv-файл и сохранить данные в excel-файл, кроме возраста -
# столбец с этими данными не нужен. Таблица должна выглядеть как на примере.
import csv
from openpyxl import Workbook
from openpyxl.styles import Alignment


def add_cell(cell: object, value: object, alignment: object):
    """
    Appends a value to the specified cell.
    """
    cell.value, cell.alignment = value, alignment


if __name__ == "__main__":
    with open('clients.csv') as file:
        reader = csv.reader(file)
        # Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "MaximGvozd"
        # Style
        align = Alignment(horizontal='center', vertical='center')
        # counters
        age_id = None
        person = column = 0
        for items in reader:
            row = 1
            if not column:
                # header of names
                row += 1
                column += 1
                for item in items:
                    if item == "age":
                        # without 'age'
                        age_id = items.index(item)
                        continue
                    add_cell(ws.cell(row=row, column=column), item, align)
                    row += 1
                continue
            column += 1
            # header of person
            person += 1
            add_cell(ws.cell(row=row, column=column), f"Person {person}", align)
            row += 1
            for item in items:
                if items.index(item) == age_id:
                    # without 'age'
                    continue
                add_cell(ws.cell(row=row, column=column), item, align)
                row += 1
        wb.save('clients.xlsx')
